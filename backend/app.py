from flask import Flask, request, jsonify, send_file, make_response, g
from dotenv import load_dotenv
from db import init_db, get_conn, row_to_dict, rows_to_list, DB_PATH, reset_sandbox, SANDBOX_MODE
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io, calendar as cal_mod, os, uuid, shutil, threading, time
from datetime import datetime
from collections import defaultdict

load_dotenv()  # loads backend/.env locally; no-op in production if env vars
                # are supplied directly by the hosting platform

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ── OCR: OCR.space ────────────────────────────────────────────────────────────
# Free OCR API — no credit card needed, 25,000 pages/month free.
# Get a free key at: https://ocr.space/ocrapi (click "Get API Key Free")
# Set OCR_SPACE_API_KEY in backend/.env locally, or as an environment
# variable on your hosting platform. Never hardcode the key here.
OCR_SPACE_API_KEY = os.environ.get("OCR_SPACE_API_KEY", "")

OCR_AVAILABLE = bool(OCR_SPACE_API_KEY)

# ── Demo mode ──────────────────────────────────────────────────────────────
# When DEMO_MODE=true (the default), the real on-disk sample data is seeded
# once and then never touched by visitors. Instead, each visitor is given
# their own private in-memory "sandbox" (see db.py) on their first request:
# every read and write they make happens in that sandbox, so they can freely
# try out the app — add a hotel, log a week's income, and so on — without
# ever changing the real demo data or affecting any other visitor. Their
# sandbox lives for as long as their browser holds the cookie (and the
# server process keeps running); there's a "Reset demo data" button in the
# UI that discards just their own sandbox. Set DEMO_MODE=false for
# local/private use if you don't want this isolation (e.g. to test the app
# exactly as a single real user would).
DEMO_MODE = os.environ.get("DEMO_MODE", "true").lower() in ("1", "true", "yes")

SANDBOX_COOKIE = "sandbox_id"

UPLOAD_DIR = os.path.join(BASE_DIR, "uploads", "packing_lists")
os.makedirs(UPLOAD_DIR, exist_ok=True)

app = Flask(__name__)

ALLOWED_ORIGIN = os.environ.get("ALLOWED_ORIGIN", "*")


@app.before_request
def assign_sandbox():
    # Every visitor gets a random id the first time they show up. db.py
    # reads this (via flask.g) to route their reads/writes into their own
    # private in-memory copy of the data instead of the real database.
    g.sandbox_id = request.cookies.get(SANDBOX_COOKIE) or uuid.uuid4().hex


@app.after_request
def cors(resp):
    resp.headers["Access-Control-Allow-Origin"]  = ALLOWED_ORIGIN
    resp.headers["Access-Control-Allow-Methods"] = "GET,POST,PATCH,DELETE,OPTIONS"
    resp.headers["Access-Control-Allow-Headers"] = "Content-Type,Authorization"
    if SANDBOX_MODE and request.cookies.get(SANDBOX_COOKIE) != getattr(g, "sandbox_id", None):
        resp.set_cookie(
            SANDBOX_COOKIE, g.sandbox_id,
            max_age=60 * 60 * 24 * 30,  # 30 days
            samesite="Lax", httponly=True,
        )
    return resp


@app.route("/", methods=["OPTIONS"])
@app.route("/<path:p>", methods=["OPTIONS"])
def options(p=""):
    return make_response("", 204)

MONTH_NAMES = ["","January","February","March","April","May","June",
               "July","August","September","October","November","December"]


# ────────────────────────────────────────────────────────────────────────────
# HOTELS
# ────────────────────────────────────────────────────────────────────────────

@app.get("/hotels")
def list_hotels():
    conn = get_conn()
    rows = rows_to_list(conn.execute("SELECT * FROM hotels ORDER BY name").fetchall())
    conn.close()
    return jsonify(rows)


@app.post("/hotels")
def create_hotel():
    d = request.json
    conn = get_conn()
    try:
        conn.execute(
            "INSERT INTO hotels (name, customer_type, active_status) VALUES (?,?,?)",
            (d["name"], d["customer_type"], d.get("active_status", 1))
        )
        conn.commit()
        hotel = row_to_dict(conn.execute("SELECT * FROM hotels WHERE name=?", (d["name"],)).fetchone())
        conn.close()
        return jsonify(hotel), 201
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 400


@app.patch("/hotels/<int:hid>")
def update_hotel(hid):
    d = request.json
    conn = get_conn()
    fields = {k: v for k, v in d.items() if k in ("name","customer_type","active_status")}
    if not fields:
        conn.close()
        return jsonify({"error": "Nothing to update"}), 400
    sql = "UPDATE hotels SET " + ", ".join(f"{k}=?" for k in fields) + " WHERE id=?"
    conn.execute(sql, list(fields.values()) + [hid])
    conn.commit()
    hotel = row_to_dict(conn.execute("SELECT * FROM hotels WHERE id=?", (hid,)).fetchone())
    conn.close()
    return jsonify(hotel)


@app.delete("/hotels/<int:hid>")
def delete_hotel(hid):
    conn = get_conn()
    conn.execute("DELETE FROM hotels WHERE id=?", (hid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# ────────────────────────────────────────────────────────────────────────────
# DAILY RESULTS  (entered every day; these add up into the monthly result)
# ────────────────────────────────────────────────────────────────────────────

def _enrich_daily(r):
    d = dict(r)
    d["total_income"] = (d["guest_laundry_income"] + d["staff_laundry_income"] +
                         d["flat_laundry_income"])
    return d


@app.get("/daily-results")
def list_daily_results():
    """List daily entries. Filter by hotel_id and/or year+month."""
    hotel_id = request.args.get("hotel_id", type=int)
    year     = request.args.get("year",     type=int)
    month    = request.args.get("month",    type=int)

    sql = """
        SELECT dr.*, h.name as hotel_name
        FROM daily_results dr
        JOIN hotels h ON h.id = dr.hotel_id
        WHERE 1=1
    """
    params = []
    if hotel_id: sql += " AND dr.hotel_id=?"; params.append(hotel_id)
    if year and month:
        last = cal_mod.monthrange(year, month)[1]
        sql += " AND dr.entry_date BETWEEN ? AND ?"
        params += [f"{year}-{month:02d}-01", f"{year}-{month:02d}-{last:02d}"]
    sql += " ORDER BY dr.entry_date"

    conn = get_conn()
    rows = [_enrich_daily(r) for r in conn.execute(sql, params).fetchall()]
    conn.close()
    return jsonify(rows)


@app.post("/daily-results")
def upsert_daily_result():
    """
    Add or update one week's figures for one hotel.
    entry_date = week start date (Monday)
    week_end_date = week end date (Sunday or last day of month)
    week_label = human-readable label e.g. "1–7 Jul"
    """
    d = request.json
    conn = get_conn()

    entry_date = d["entry_date"]
    year, month = int(entry_date[:4]), int(entry_date[5:7])
    status_row = row_to_dict(conn.execute(
        "SELECT * FROM monthly_status WHERE hotel_id=? AND year=? AND month=?",
        (d["hotel_id"], year, month)
    ).fetchone())
    if status_row and status_row["status"] == "finalised" and not d.get("correction_note"):
        conn.close()
        return jsonify({"error": "This month is finalised. A correction note is required to make changes."}), 400

    try:
        conn.execute("""
            INSERT INTO daily_results
              (hotel_id, entry_date, week_end_date, week_label,
               guest_laundry_income, staff_laundry_income,
               flat_laundry_income, notes)
            VALUES (?,?,?,?,?,?,?,?)
            ON CONFLICT(hotel_id, entry_date) DO UPDATE SET
              week_end_date=excluded.week_end_date,
              week_label=excluded.week_label,
              guest_laundry_income=excluded.guest_laundry_income,
              staff_laundry_income=excluded.staff_laundry_income,
              flat_laundry_income=excluded.flat_laundry_income,
              notes=excluded.notes
        """, (
            d["hotel_id"], entry_date,
            d.get("week_end_date", entry_date),
            d.get("week_label", ""),
            d.get("guest_laundry_income", 0),
            d.get("staff_laundry_income", 0),
            d.get("flat_laundry_income", 0),
            d.get("notes", ""),
        ))
        conn.commit()
        row = conn.execute("""
            SELECT dr.*, h.name as hotel_name FROM daily_results dr
            JOIN hotels h ON h.id=dr.hotel_id
            WHERE dr.hotel_id=? AND dr.entry_date=?
        """, (d["hotel_id"], entry_date)).fetchone()
        conn.close()
        return jsonify(_enrich_daily(row)), 201
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 400


@app.delete("/daily-results/<int:rid>")
def delete_daily_result(rid):
    conn = get_conn()
    conn.execute("DELETE FROM daily_results WHERE id=?", (rid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.get("/weekly-schedule/<int:year>/<int:month>")
def weekly_schedule(year, month):
    """
    Returns the auto-generated week ranges for a given month.
    Weeks run Monday–Sunday, split at month boundaries.
    The frontend uses this to pre-populate the weekly entry grid,
    and the manager can adjust any week's start/end date if needed.
    """
    from datetime import date, timedelta
    import calendar as cal

    first_day = date(year, month, 1)
    last_day  = date(year, month, cal.monthrange(year, month)[1])

    weeks = []
    current = first_day
    week_num = 1

    while current <= last_day:
        # Week ends on the coming Sunday, or last day of month, whichever is first
        days_until_sunday = 6 - current.weekday()  # 0=Mon…6=Sun
        week_end = min(current + timedelta(days=days_until_sunday), last_day)

        start_str = current.strftime("%Y-%m-%d")
        end_str   = week_end.strftime("%Y-%m-%d")

        # Label e.g. "1–7 Jul" or "28 Jul–3 Aug" (cross-month weeks capped at month boundary)
        if current.month == week_end.month:
            label = f"{current.day}–{week_end.day} {current.strftime('%b')}"
        else:
            label = f"{current.day} {current.strftime('%b')}–{week_end.day} {week_end.strftime('%b')}"

        weeks.append({
            "week_num": week_num,
            "week_start": start_str,
            "week_end":   end_str,
            "label":      label,
        })
        current = week_end + timedelta(days=1)
        week_num += 1

    return jsonify({"year": year, "month": month, "weeks": weeks})


# ────────────────────────────────────────────────────────────────────────────
# MONTHLY STATUS  (in_progress / ready_for_checking / checked / finalised)
# ────────────────────────────────────────────────────────────────────────────

@app.get("/monthly-status/<int:year>/<int:month>")
def list_monthly_status(year, month):
    conn = get_conn()
    rows = rows_to_list(conn.execute("""
        SELECT ms.*, h.name as hotel_name FROM monthly_status ms
        JOIN hotels h ON h.id=ms.hotel_id
        WHERE ms.year=? AND ms.month=?
    """, (year, month)).fetchall())
    conn.close()
    return jsonify(rows)


@app.post("/monthly-status")
def set_monthly_status():
    """
    Changes a month's status for a hotel. Un-finalising is a simple one-click
    action with no note required — the correction-note requirement only
    applies to editing the actual daily figures of a finalised month
    (see /daily-results), which is where the protection actually matters.
    """
    d = request.json
    conn = get_conn()
    conn.execute("""
        INSERT INTO monthly_status (hotel_id, year, month, status, correction_note)
        VALUES (?,?,?,?,?)
        ON CONFLICT(hotel_id, year, month) DO UPDATE SET
          status=excluded.status, correction_note=excluded.correction_note
    """, (d["hotel_id"], d["year"], d["month"], d.get("status","in_progress"), d.get("correction_note","")))
    conn.commit()
    row = row_to_dict(conn.execute(
        "SELECT * FROM monthly_status WHERE hotel_id=? AND year=? AND month=?",
        (d["hotel_id"], d["year"], d["month"])
    ).fetchone())
    conn.close()
    return jsonify(row), 201


# ────────────────────────────────────────────────────────────────────────────
# MONTHLY RESULTS  (calculated automatically — the sum of all daily entries)
# ────────────────────────────────────────────────────────────────────────────

@app.get("/monthly-results")
def list_results():
    """Rolled-up monthly totals per hotel — calculated from daily_results."""
    year = request.args.get("year", type=int)
    month = request.args.get("month", type=int)

    conn = get_conn()
    hotels = rows_to_list(conn.execute("SELECT * FROM hotels ORDER BY name").fetchall())
    last = cal_mod.monthrange(year, month)[1] if year and month else None

    rows = []
    for h in hotels:
        sql = """
            SELECT
              COALESCE(SUM(guest_laundry_income),0) as guest_laundry_income,
              COALESCE(SUM(staff_laundry_income),0) as staff_laundry_income,
              COALESCE(SUM(flat_laundry_income),0)  as flat_laundry_income,
              COUNT(*) as days_entered
            FROM daily_results
            WHERE hotel_id=? AND entry_date BETWEEN ? AND ?
        """
        totals = row_to_dict(conn.execute(sql, (
            h["id"], f"{year}-{month:02d}-01", f"{year}-{month:02d}-{last:02d}"
        )).fetchone())

        status_row = row_to_dict(conn.execute(
            "SELECT * FROM monthly_status WHERE hotel_id=? AND year=? AND month=?",
            (h["id"], year, month)
        ).fetchone())

        total_income = (totals["guest_laundry_income"] + totals["staff_laundry_income"] +
                        totals["flat_laundry_income"])
        rows.append({
            "hotel_id": h["id"], "hotel_name": h["name"],
            "year": year, "month": month,
            "guest_laundry_income": totals["guest_laundry_income"],
            "staff_laundry_income": totals["staff_laundry_income"],
            "flat_laundry_income":  totals["flat_laundry_income"],
            "total_income": total_income,
            "days_entered": totals["days_entered"],
            "status": status_row["status"] if status_row else "in_progress",
            "correction_note": status_row["correction_note"] if status_row else "",
        })
    conn.close()
    return jsonify(rows)


@app.get("/monthly-results/summary/<int:year>/<int:month>")
def monthly_summary(year, month):
    conn = get_conn()
    last = cal_mod.monthrange(year, month)[1]

    active_hotels = rows_to_list(conn.execute("SELECT * FROM hotels WHERE active_status=1").fetchall())
    daily = rows_to_list(conn.execute("""
        SELECT dr.*, h.name as hotel_name, h.active_status FROM daily_results dr
        JOIN hotels h ON h.id=dr.hotel_id
        WHERE dr.entry_date BETWEEN ? AND ?
    """, (f"{year}-{month:02d}-01", f"{year}-{month:02d}-{last:02d}")).fetchall())
    conn.close()

    by_hotel = defaultdict(lambda: {"guest":0,"staff":0,"flat":0,"days":0})
    for r in daily:
        b = by_hotel[r["hotel_id"]]
        b["guest"] += r["guest_laundry_income"]
        b["staff"] += r["staff_laundry_income"]
        b["flat"]  += r["flat_laundry_income"]
        b["days"]  += 1

    hotel_ids_with_data = set(by_hotel.keys())
    missing = [h["name"] for h in active_hotels if h["id"] not in hotel_ids_with_data]

    rows = []
    for hid, b in by_hotel.items():
        name = next((r["hotel_name"] for r in daily if r["hotel_id"]==hid), "Unknown")
        rows.append({
            "hotel_id": hid, "hotel_name": name,
            "guest_laundry_income": b["guest"], "staff_laundry_income": b["staff"],
            "flat_laundry_income": b["flat"],
            "total_income": b["guest"]+b["staff"]+b["flat"],
            "days_entered": b["days"],
        })

    return jsonify({
        "year": year, "month": month,
        "grand_total": sum(r["total_income"] for r in rows),
        "total_guest":        sum(r["guest_laundry_income"] for r in rows),
        "total_staff":        sum(r["staff_laundry_income"] for r in rows),
        "total_flat":         sum(r["flat_laundry_income"]  for r in rows),
        "hotel_count": len(rows),
        "missing_hotels": missing,
        "rows": rows,
    })


@app.get("/monthly-results/comparison/<int:year>/<int:month>")
def comparison(year, month):
    prev_year  = year if month > 1 else year - 1
    prev_month = month - 1 if month > 1 else 12

    conn = get_conn()
    def fetch_totals(y, m):
        last = cal_mod.monthrange(y, m)[1]
        rows = conn.execute("""
            SELECT dr.*, h.name as hotel_name FROM daily_results dr
            JOIN hotels h ON h.id=dr.hotel_id
            WHERE dr.entry_date BETWEEN ? AND ?
        """, (f"{y}-{m:02d}-01", f"{y}-{m:02d}-{last:02d}")).fetchall()
        totals = defaultdict(lambda: {"guest_laundry_income":0,"staff_laundry_income":0,
                                       "flat_laundry_income":0,"hotel_name":""})
        for r in rows:
            t = totals[r["hotel_id"]]
            t["guest_laundry_income"] += r["guest_laundry_income"]
            t["staff_laundry_income"] += r["staff_laundry_income"]
            t["flat_laundry_income"]  += r["flat_laundry_income"]
            t["hotel_name"] = r["hotel_name"]
        return totals

    cur  = fetch_totals(year,  month)
    prev = fetch_totals(prev_year, prev_month)
    hotels_map = {h["id"]: h["name"] for h in rows_to_list(conn.execute("SELECT * FROM hotels").fetchall())}
    conn.close()

    def total_of(t): return t["guest_laundry_income"]+t["staff_laundry_income"]+t["flat_laundry_income"]

    all_ids = set(cur) | set(prev)
    rows = []
    for hid in sorted(all_ids):
        ct = total_of(cur.get(hid, {"guest_laundry_income":0,"staff_laundry_income":0,"flat_laundry_income":0}))
        pt = total_of(prev.get(hid, {"guest_laundry_income":0,"staff_laundry_income":0,"flat_laundry_income":0}))
        chg = ct - pt
        pct = round(chg / pt * 100, 1) if pt else None
        rows.append({
            "hotel_id": hid, "hotel_name": hotels_map.get(hid,"?"),
            "current_total": ct, "previous_total": pt,
            "change": chg, "pct_change": pct
        })

    with_prev = [r for r in rows if r["previous_total"] > 0]
    services = {}
    for field, label in [("guest_laundry_income","Guest Laundry"),
                          ("staff_laundry_income","Staff Laundry"),
                          ("flat_laundry_income","Flat Laundry")]:
        ct = sum(v[field] for v in cur.values())
        pt = sum(v[field] for v in prev.values())
        diff = ct - pt
        services[label] = {"current": ct, "previous": pt, "change": diff,
                           "pct_change": round(diff/pt*100,1) if pt else None}

    return jsonify({
        "current_period":  {"year": year,      "month": month},
        "previous_period": {"year": prev_year, "month": prev_month},
        "hotel_rows": rows,
        "service_comparison": services,
        "biggest_increase": max(with_prev, key=lambda r: r["change"], default=None),
        "biggest_decrease": min(with_prev, key=lambda r: r["change"], default=None),
    })


@app.get("/compare")
def flexible_compare():
    """
    Free-form comparison: pick any hotel (or 'all'), and any two periods
    (not necessarily consecutive months). Used by the Comparison page's
    preset question buttons.

    Query params:
      hotel_id   - a hotel id, or omit/'' for all hotels combined
      year_a, month_a - first period
      year_b, month_b - second period
    """
    hotel_id = request.args.get("hotel_id", type=int)  # None = all hotels
    year_a  = request.args.get("year_a",  type=int)
    month_a = request.args.get("month_a", type=int)
    year_b  = request.args.get("year_b",  type=int)
    month_b = request.args.get("month_b", type=int)

    if not all([year_a, month_a, year_b, month_b]):
        return jsonify({"error": "year_a, month_a, year_b, and month_b are all required."}), 400

    conn = get_conn()

    def period_totals(y, m):
        """Returns {hotel_id: {hotel_name, total_income, guest/staff/flat}} for one period."""
        all_totals = _monthly_totals(conn, y, m)
        if hotel_id:
            return {hid: t for hid, t in all_totals.items() if hid == hotel_id}
        return all_totals

    totals_a = period_totals(year_a, month_a)
    totals_b = period_totals(year_b, month_b)
    conn.close()

    def grand(totals): return sum(t["total_income"] for t in totals.values())

    grand_a = grand(totals_a)
    grand_b = grand(totals_b)
    change = grand_b - grand_a
    pct = round(change / grand_a * 100, 1) if grand_a else None

    # Per-hotel breakdown (relevant when comparing "all hotels", or just the one hotel chosen)
    all_ids = set(totals_a) | set(totals_b)
    hotel_rows = []
    for hid in all_ids:
        ta = totals_a.get(hid, {"hotel_name": "Unknown", "total_income": 0})
        tb = totals_b.get(hid, {"hotel_name": "Unknown", "total_income": 0})
        name = ta.get("hotel_name") or tb.get("hotel_name")
        a_total = ta["total_income"]
        b_total = tb["total_income"]
        diff = b_total - a_total
        hotel_rows.append({
            "hotel_id": hid, "hotel_name": name,
            "period_a_total": a_total, "period_b_total": b_total,
            "change": diff,
            "pct_change": round(diff / a_total * 100, 1) if a_total else None,
        })
    hotel_rows.sort(key=lambda r: r["hotel_name"])

    with_base = [r for r in hotel_rows if r["period_a_total"] > 0]
    biggest_increase = max(with_base, key=lambda r: r["change"], default=None)
    biggest_decrease = min(with_base, key=lambda r: r["change"], default=None)

    return jsonify({
        "period_a": {"year": year_a, "month": month_a, "label": f"{MONTH_NAMES[month_a]} {year_a}"},
        "period_b": {"year": year_b, "month": month_b, "label": f"{MONTH_NAMES[month_b]} {year_b}"},
        "scope": "single_hotel" if hotel_id else "all_hotels",
        "grand_total_a": grand_a,
        "grand_total_b": grand_b,
        "change": change,
        "pct_change": pct,
        "hotel_rows": hotel_rows,
        "biggest_increase": biggest_increase,
        "biggest_decrease": biggest_decrease,
    })


# ────────────────────────────────────────────────────────────────────────────
# POOL STOCK
# ────────────────────────────────────────────────────────────────────────────

@app.get("/pool-stock")
def list_pool_stock():
    hotel_id   = request.args.get("hotel_id",   type=int)
    year       = request.args.get("year",        type=int)
    month      = request.args.get("month",       type=int)
    linen      = request.args.get("linen_item")
    ref        = request.args.get("packing_list_ref")
    entry_date = request.args.get("entry_date")  # exact date filter e.g. "2025-03-07"

    sql = """SELECT ps.*, h.name as hotel_name FROM pool_stock ps
             JOIN hotels h ON h.id=ps.hotel_id WHERE 1=1"""
    params = []
    if hotel_id:   sql += " AND ps.hotel_id=?";              params.append(hotel_id)
    if entry_date: sql += " AND ps.entry_date=?";            params.append(entry_date)
    elif year and month:
        last = cal_mod.monthrange(year, month)[1]
        sql += " AND ps.entry_date BETWEEN ? AND ?"
        params += [f"{year}-{month:02d}-01", f"{year}-{month:02d}-{last:02d}"]
    elif year:
        sql += " AND ps.entry_date BETWEEN ? AND ?"
        params += [f"{year}-01-01", f"{year}-12-31"]
    if linen: sql += " AND ps.linen_item LIKE ?"; params.append(f"%{linen}%")
    if ref:   sql += " AND ps.packing_list_ref LIKE ?"; params.append(f"%{ref}%")
    sql += " ORDER BY ps.entry_date DESC, h.name"

    conn = get_conn()
    rows = rows_to_list(conn.execute(sql, params).fetchall())
    conn.close()
    return jsonify(rows)


@app.post("/pool-stock")
def create_pool_stock():
    d = request.json
    conn = get_conn()
    # Duplicate ref check
    if d.get("packing_list_ref"):
        dup = conn.execute("""
            SELECT id FROM pool_stock
            WHERE hotel_id=? AND packing_list_ref=? AND linen_item=?
        """, (d["hotel_id"], d["packing_list_ref"], d["linen_item"])).fetchone()
        if dup:
            conn.close()
            return jsonify({"error": f"Duplicate packing list reference '{d['packing_list_ref']}' for this hotel and item."}), 400
    try:
        conn.execute("""
            INSERT INTO pool_stock (hotel_id, entry_date, linen_item, quantity, packing_list_ref, notes)
            VALUES (?,?,?,?,?,?)
        """, (d["hotel_id"], d["entry_date"], d["linen_item"],
              d["quantity"], d.get("packing_list_ref",""), d.get("notes","")))
        conn.commit()
        eid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        row = row_to_dict(conn.execute("""
            SELECT ps.*, h.name as hotel_name FROM pool_stock ps
            JOIN hotels h ON h.id=ps.hotel_id WHERE ps.id=?
        """, (eid,)).fetchone())
        conn.close()
        return jsonify(row), 201
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 400


@app.patch("/pool-stock/<int:eid>")
def update_pool_stock(eid):
    d = request.json
    conn = get_conn()
    allowed = ["entry_date","linen_item","quantity","packing_list_ref","notes"]
    fields = {k: v for k, v in d.items() if k in allowed}
    sql = "UPDATE pool_stock SET " + ", ".join(f"{k}=?" for k in fields) + " WHERE id=?"
    conn.execute(sql, list(fields.values()) + [eid])
    conn.commit()
    row = row_to_dict(conn.execute("""
        SELECT ps.*, h.name as hotel_name FROM pool_stock ps
        JOIN hotels h ON h.id=ps.hotel_id WHERE ps.id=?
    """, (eid,)).fetchone())
    conn.close()
    return jsonify(row)


@app.delete("/pool-stock/<int:eid>")
def delete_pool_stock(eid):
    conn = get_conn()
    conn.execute("DELETE FROM pool_stock WHERE id=?", (eid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.get("/pool-stock/summary/<int:year>/<int:month>")
def pool_stock_summary(year, month):
    last = cal_mod.monthrange(year, month)[1]
    conn = get_conn()
    entries = rows_to_list(conn.execute("""
        SELECT ps.*, h.name as hotel_name FROM pool_stock ps
        JOIN hotels h ON h.id=ps.hotel_id
        WHERE ps.entry_date BETWEEN ? AND ?
    """, (f"{year}-{month:02d}-01", f"{year}-{month:02d}-{last:02d}")).fetchall())
    conn.close()

    by_hotel = defaultdict(int)
    by_item  = defaultdict(int)
    by_date  = defaultdict(int)
    ref_map  = defaultdict(list)

    for e in entries:
        by_hotel[e["hotel_name"]] += e["quantity"]
        by_item[e["linen_item"]]  += e["quantity"]
        by_date[e["entry_date"]]  += e["quantity"]
        if e["packing_list_ref"]:
            ref_map[e["packing_list_ref"]].append(e["id"])

    dup_refs = {k: v for k, v in ref_map.items() if len(v) > 1}

    return jsonify({
        "year": year, "month": month,
        "overall_total": sum(e["quantity"] for e in entries),
        "by_hotel":      dict(sorted(by_hotel.items())),
        "by_linen_item": dict(sorted(by_item.items())),
        "by_date":       dict(sorted(by_date.items())),
        "entry_count":   len(entries),
        "checks": {
            "duplicate_packing_list_refs": dup_refs
        }
    })


# ────────────────────────────────────────────────────────────────────────────
# PACKING LIST SCANNER  (photo upload → OCR → review screen → confirm & save)
# ────────────────────────────────────────────────────────────────────────────

ALLOWED_EXTENSIONS = {"jpg","jpeg","png","heic"}

# Per-hotel linen item templates — exact order from each hotel's printed form.
# When a packing list is scanned, the system uses these to pre-fill item names
# automatically so the reviewer only needs to check/correct the numbers.
# Replace the keys below with your actual customer names to enable template matching.

HOTEL_TEMPLATES = {
    "Hotel A": [
        "Single Sheets","Queen Sheets","King Sheets",
        "Single Duvet Covers","Double Duvet Covers","Emperor King Duvet Covers",
        "Queen Pillow Cases","King Pillow Cases",
        "Scallop King Bedspreads","Escallop Single Bedspreads","Queen Bedspreads",
        "Hand Towels","Bath Towels","Bath Sheets","Bath Mats","Face Cloths",
        "Bathrobes - Green",
        "Small Table Cloths 54x54","Round Table Cloths","Large Table Cloths 225x235",
        "FH Napkins","Table Runners","Events Large Table Runners",
        "Small Tray Cloths","Large Tray Cloths","Glass Cloths","Oven Cloths",
    ],
    "Hotel B": [
        "Single Sheets","Single Duvet Covers","Emperor King Duvet Covers",
        "Double Duvet Covers","King Sheets",
        "Queen Pillow Cases","Single Bedspreads","King Bedspreads",
        "Hand Towels","Bath Towels","Bath Sheets","Bath Mats","Face Cloths",
        "Spa Jumbo Bath Sheets","Bathrobes",
        "Ferguson Table Cloths 54X54","Ferguson Napkins",
        "Old Style Table Napkins","Old Style Table Cloth 54X54",
        "Large Table Cloths 225X235","Round Table Cloths","FH Napkins",
        "Small Tray Cloths","Large Tray Cloths",
        "Ferguson Round Table Cloths","Medium Table Cloths 180X270",
        "Glass Cloths","Oven Cloths",
    ],
    "Hotel C": [
        "Single Sheets","Queen Sheets","King Sheets",
        "Single Duvet Covers","Queen Duvet Covers","Emperor Duvet Covers",
        "Double Duvet Covers","Queen Pillow Cases","King Pillow Cases",
        "Single Bedspreads","Queen Bedspreads","King Bedspreads",
        "Hand Towels","Bath Towels","Bath Sheets","Bath Mats","Face Cloths",
        "Bathrobes",
        "Small Table Cloths 54X54","Medium Table Cloths 180X270",
        "Large Round Table Cloths","Table Napkins","FH Napkins","Table Runners",
        "Small Tray Cloths","Large Tray Cloths","Glass Cloths","Oven Cloths",
    ],
    "Hotel D": [
        "Single Sheets","Queen Sheets","King Sheets",
        "Single Duvet Covers","Double Duvet Covers",
        "Queen Duvet Covers","Emperor Duvet Covers",
        "Queen Pillow Cases","King Pillow Cases",
        "Single Bedspreads","Queen Bedspreads","King Bedspreads",
        "Hand Towels","Bath Towels","Bath Sheets","Bath Mats","Face Cloths",
        "Bathrobes - Orange",
        "Small Table Cloths 54X54","Medium Table Cloths 180X270",
        "Large Round Table Cloths","Large Table Cloths 225X235",
        "FH Napkins","Table Runners",
        "Small Tray Cloths","Large Tray Cloths",
        "Glass Cloths","Kitchen Cloths","Oven Cloths",
    ],
    "Hotel E": [
        "Single Sheets","Queen Sheets","King Sheets",
        "Single Duvet Covers","Double Duvet Covers",
        "King Duvet Covers","Emperor Duvet Covers",
        "Queen Pillow Cases","King Pillow Cases",
        "Single Bedspreads","Queen Bedspreads","King Bedspreads",
        "Hand Towels","Bath Towels","Bath Sheets","Bath Mats","Face Cloths",
        "Bathrobes - Red",
        "Ferguson Table Cloths 54X54","Ferguson Table Cloths 108X108",
        "Ferguson Napkins (new)","Ferguson Napkins (old)",
        "Round Table Cloths","Large Table Cloths 225X235","FH Napkins",
        "Table Runners","Small Tray Cloths","Glass Cloths","Oven Cloths",
    ],
    "Hotel F": [
        "Single Sheets","Queen Sheets","King Sheets",
        "Single Duvet Covers","Queen Duvet Covers",
        "King Duvet Covers","Emperor Duvet Covers",
        "Double Duvet Covers (for twins)",
        "King Pillow Cases","Queen Pillow Cases",
        "Single Bedspreads","Queen Bedspreads","King Bedspreads",
        "Super King Bedspreads",
        "Hand Towels","Bath Towels","Bath Sheets",
        "Bath Mats","Big Bath Mats","Face Cloths",
        "Bathrobes - Red",
        "FH Napkins",
        "Small Tray Cloths","Large Tray Cloths","Table Runners",
        "Kitchen Cloths","Oven Cloths",
    ],
    "Hotel G": [
        "Single Sheets","Queen Sheets","King Sheets",
        "Single Duvet Covers","Double Duvet Covers",
        "Queen Duvet Covers","Emperor Duvet Covers",
        "Queen Duvet Covers (Twin beds)",
        "King Pillow Cases",
        "Single Bedspreads","Queen Bedspreads","King Bedspreads",
        "Extra King Bedspreads",
        "Hand Towels","Bath Towels","Face Cloths",
        "Bath Mats","Bath Sheets",
        "Bathrobes - Blue Velour",
        "Satin Band Table Cloths 54X54","Large Table Cloths 54X54",
        "Large Round Table Cloths","Large Table Cloths 225X235",
        "FH Napkins","Pure Linen Napkins","Table Runners",
        "Small Tray Cloths","Large Tray Cloths",
        "Glass Cloths","Kitchen Cloths","Oven Cloths",
    ],
    "Hotel H": [
        "Single Sheets","Queen Sheets","King Sheets",
        "Single Duvet Covers","Double Duvet Covers","Emperor King Duvet Covers",
        "Queen Pillow Cases","King Pillow Cases",
        "Single Bedspreads","Queen Bedspreads","King Bedspreads",
        "Hand Towels","Bath Towels","Bath Sheets","Bath Mats","Face Cloths",
        "Bathrobes - Green",
        "Small Table Cloths 54X54","Medium Table Cloths 180X270",
        "Large Round Table Cloths","Large Table Cloths 225X235",
        "FH Napkins","Table Runners",
        "Small Tray Cloths","Large Tray Cloths","Glass Cloths","Oven Cloths",
    ],
}

LINEN_ITEMS_GENERIC = [
    "Single Sheets","Queen Sheets","King Sheets",
    "Single Duvet Covers","Double Duvet Covers","Queen Duvet Covers",
    "Emperor Duvet Covers","Queen Pillow Cases","King Pillow Cases",
    "Single Bedspreads","Queen Bedspreads","King Bedspreads",
    "Hand Towels","Bath Towels","Bath Sheets","Bath Mats","Face Cloths",
    "Bathrobes","Table Cloths","Napkins","Tray Cloths","Glass Cloths","Oven Cloths",
]


def _get_hotel_template(hotel_name):
    for key in HOTEL_TEMPLATES:
        if key.lower() in hotel_name.lower() or hotel_name.lower() in key.lower():
            return HOTEL_TEMPLATES[key]
    return LINEN_ITEMS_GENERIC


def _allowed_file(filename):
    return "." in filename and filename.rsplit(".",1)[1].lower() in ALLOWED_EXTENSIONS


def _run_ocr(image_path):
    """
    Uses OCR.space API to extract numbers from a packing list photo.
    Free tier: 25,000 pages/month, no credit card needed.
    Every result MUST be reviewed by a human before saving.
    """
    if not OCR_AVAILABLE:
        return "", []

    import base64, urllib.request, urllib.parse, urllib.error, json as _json, re

    # Read and base64-encode the image
    with open(image_path, "rb") as f:
        image_data = base64.b64encode(f.read()).decode("utf-8")

    ext = image_path.rsplit(".", 1)[-1].lower()
    mime = {"jpg":"image/jpeg","jpeg":"image/jpeg",
            "png":"image/png","heic":"image/heic"}.get(ext, "image/jpeg")

    # OCR.space API call
    # isTable=true helps with structured forms with rows and columns
    post_data = urllib.parse.urlencode({
        "apikey":          OCR_SPACE_API_KEY,
        "base64Image":     f"data:{mime};base64,{image_data}",
        "language":        "eng",
        "isOverlayRequired": "false",
        "detectOrientation": "true",
        "scale":           "true",
        "isTable":         "true",
        "OCREngine":       "2",   # Engine 2 is better for printed text and tables
    }).encode("utf-8")

    req = urllib.request.Request(
        "https://api.ocr.space/parse/image",
        data=post_data,
        headers={"Content-Type": "application/x-www-form-urlencoded"}
    )

    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            result = _json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        return f"API error: {e.read().decode()}", []
    except Exception as e:
        return f"Network error: {str(e)}", []

    if result.get("IsErroredOnProcessing"):
        return f"OCR error: {result.get('ErrorMessage','unknown')}", []

    parsed_results = result.get("ParsedResults", [])
    if not parsed_results:
        return "", []

    raw_text = parsed_results[0].get("ParsedText", "")

    # Extract numbers from each line
    detected = []
    for line in raw_text.split("\n"):
        line = line.strip()
        if not line:
            continue
        numbers = re.findall(r"\b\d{1,4}\b", line)
        if numbers:
            qty = int(numbers[-1])
            label_guess = re.sub(r"\d+", "", line).strip(" .:-|\t")
            detected.append({
                "raw_line":      line,
                "label_guess":   label_guess or "Unrecognised",
                "quantity_guess": qty
            })

    return raw_text, detected


@app.post("/pool-stock/scan")
def scan_packing_list():
    """
    Step 1: upload a photo of a packing list. Runs OCR and returns a
    best-guess list of items/quantities for the human to review — nothing
    is saved to pool_stock yet.
    """
    if "image" not in request.files:
        return jsonify({"error": "No image file was uploaded."}), 400
    file = request.files["image"]
    hotel_id   = request.form.get("hotel_id", type=int)
    entry_date = request.form.get("entry_date")

    if not hotel_id or not entry_date:
        return jsonify({"error": "Hotel and date are required."}), 400
    if file.filename == "" or not _allowed_file(file.filename):
        return jsonify({"error": "Please upload a JPG, PNG, or HEIC image."}), 400

    ext = file.filename.rsplit(".",1)[1].lower()
    saved_name = f"{uuid.uuid4().hex}.{ext}"
    saved_path = os.path.join(UPLOAD_DIR, saved_name)
    file.save(saved_path)

    raw_text, detected = _run_ocr(saved_path)

    # Look up hotel name so we can use the right template
    conn = get_conn()
    hotel = row_to_dict(conn.execute("SELECT * FROM hotels WHERE id=?", (hotel_id,)).fetchone())
    hotel_name = hotel["name"] if hotel else ""
    template = _get_hotel_template(hotel_name)

    # Build pre-filled rows using the hotel's exact item order.
    # We try to match OCR-detected numbers to template positions in order —
    # since the items are in a fixed known order, the nth number found likely
    # belongs to the nth item. This is much more reliable than trying to match
    # by item name from messy OCR text.
    ocr_numbers = [line["quantity_guess"] for line in detected] if detected else []

    pre_filled_rows = []
    for i, item_name in enumerate(template):
        qty = ocr_numbers[i] if i < len(ocr_numbers) else ""
        pre_filled_rows.append({
            "linen_item": item_name,
            "quantity": qty,
            "ocr_confidence": "auto" if i < len(ocr_numbers) else "manual",
        })

    conn.execute("""
        INSERT INTO scanned_packing_lists (hotel_id, entry_date, image_filename, raw_ocr_text, status)
        VALUES (?,?,?,?, 'pending_review')
    """, (hotel_id, entry_date, saved_name, raw_text))
    conn.commit()
    scan_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.close()

    using_template = hotel_name in " ".join(HOTEL_TEMPLATES.keys()) or any(
        k.lower() in hotel_name.lower() for k in HOTEL_TEMPLATES
    )

    return jsonify({
        "scan_id": scan_id,
        "hotel_name": hotel_name,
        "using_hotel_template": using_template,
        "ocr_available": OCR_AVAILABLE,
        "pre_filled_rows": pre_filled_rows,
        "template_item_count": len(template),
        "ocr_numbers_found": len(ocr_numbers),
        "warning": None if OCR_AVAILABLE else
            "OCR is not available — quantities have been left blank. Fill them in manually below.",
    }), 201


@app.get("/pool-stock/scan/<int:scan_id>/image")
def get_scan_image(scan_id):
    conn = get_conn()
    row = row_to_dict(conn.execute("SELECT * FROM scanned_packing_lists WHERE id=?", (scan_id,)).fetchone())
    conn.close()
    if not row:
        return jsonify({"error": "Scan not found."}), 404
    return send_file(os.path.join(UPLOAD_DIR, row["image_filename"]))


@app.post("/pool-stock/scan/<int:scan_id>/confirm")
def confirm_scan(scan_id):
    """
    Step 2: after the human has reviewed/corrected the detected lines,
    this saves the final confirmed list into pool_stock and marks the
    scan as confirmed.
    """
    d = request.json
    conn = get_conn()
    scan = row_to_dict(conn.execute("SELECT * FROM scanned_packing_lists WHERE id=?", (scan_id,)).fetchone())
    if not scan:
        conn.close()
        return jsonify({"error": "Scan not found."}), 404

    items = d.get("items", [])
    packing_list_ref = d.get("packing_list_ref", f"SCAN-{scan_id}")
    saved_count = 0
    for item in items:
        if not item.get("linen_item") or not item.get("quantity"):
            continue
        conn.execute("""
            INSERT INTO pool_stock (hotel_id, entry_date, linen_item, quantity, packing_list_ref, notes)
            VALUES (?,?,?,?,?,?)
        """, (
            scan["hotel_id"], scan["entry_date"], item["linen_item"],
            int(item["quantity"]), packing_list_ref,
            "Added via photo scan — reviewed and confirmed by staff."
        ))
        saved_count += 1

    conn.execute("UPDATE scanned_packing_lists SET status='confirmed' WHERE id=?", (scan_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "items_saved": saved_count})


@app.post("/pool-stock/scan/<int:scan_id>/discard")
def discard_scan(scan_id):
    conn = get_conn()
    conn.execute("UPDATE scanned_packing_lists SET status='discarded' WHERE id=?", (scan_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.get("/pool-stock/scans")
def list_scans():
    """Recent scans, useful for an 'awaiting review' inbox view."""
    status = request.args.get("status")
    conn = get_conn()
    sql = """
        SELECT spl.*, h.name as hotel_name FROM scanned_packing_lists spl
        JOIN hotels h ON h.id=spl.hotel_id WHERE 1=1
    """
    params = []
    if status:
        sql += " AND spl.status=?"; params.append(status)
    sql += " ORDER BY spl.created_at DESC LIMIT 50"
    rows = rows_to_list(conn.execute(sql, params).fetchall())
    conn.close()
    return jsonify(rows)


# ────────────────────────────────────────────────────────────────────────────
# ANALYTICS
# ────────────────────────────────────────────────────────────────────────────

def _monthly_totals(conn, year, month):
    """Returns {hotel_id: {hotel_name, guest_laundry_income, ..., total_income, days_entered}}"""
    last = cal_mod.monthrange(year, month)[1]
    rows = conn.execute("""
        SELECT dr.*, h.name as hotel_name FROM daily_results dr
        JOIN hotels h ON h.id=dr.hotel_id
        WHERE dr.entry_date BETWEEN ? AND ?
    """, (f"{year}-{month:02d}-01", f"{year}-{month:02d}-{last:02d}")).fetchall()

    totals = {}
    for r in rows:
        hid = r["hotel_id"]
        if hid not in totals:
            totals[hid] = {"hotel_id": hid, "hotel_name": r["hotel_name"],
                           "guest_laundry_income":0,"staff_laundry_income":0,
                           "flat_laundry_income":0,"days_entered":0}
        totals[hid]["guest_laundry_income"] += r["guest_laundry_income"]
        totals[hid]["staff_laundry_income"] += r["staff_laundry_income"]
        totals[hid]["flat_laundry_income"]  += r["flat_laundry_income"]
        totals[hid]["days_entered"] += 1
    for t in totals.values():
        t["total_income"] = (t["guest_laundry_income"] + t["staff_laundry_income"] +
                             t["flat_laundry_income"])
    return totals


@app.get("/analytics/ranking/<int:year>/<int:month>")
def hotel_ranking(year, month):
    prev_year  = year if month > 1 else year - 1
    prev_month = month - 1 if month > 1 else 12
    last = cal_mod.monthrange(year, month)[1]

    conn = get_conn()
    current  = _monthly_totals(conn, year, month)
    previous = _monthly_totals(conn, prev_year, prev_month)
    stock = conn.execute("""
        SELECT hotel_id, SUM(quantity) as vol FROM pool_stock
        WHERE entry_date BETWEEN ? AND ? GROUP BY hotel_id
    """, (f"{year}-{month:02d}-01", f"{year}-{month:02d}-{last:02d}")).fetchall()
    stock_vol = {r["hotel_id"]: r["vol"] for r in stock}
    conn.close()

    grand = sum(r["total_income"] for r in current.values())
    rows = []
    for hid, r in current.items():
        total = r["total_income"]
        pt = previous.get(hid, {}).get("total_income", 0)
        growth = round((total - pt) / pt * 100, 1) if pt else None
        contrib = round(total / grand * 100, 1) if grand else 0
        if growth is None:  label = "New"
        elif growth >= 5:   label = "Strong"
        elif growth >= -5:  label = "Stable"
        else:               label = "Needs Attention"
        rows.append({
            "hotel_id": hid, "hotel_name": r["hotel_name"],
            "total_revenue": total, "previous_revenue": pt,
            "revenue_growth_pct": growth, "pool_stock_volume": stock_vol.get(hid,0),
            "contribution_pct": contrib, "label": label,
        })

    rows.sort(key=lambda x: x["total_revenue"], reverse=True)
    for i, row in enumerate(rows): row["rank"] = i + 1
    return jsonify({"year": year, "month": month, "grand_total": grand, "rankings": rows})


@app.get("/analytics/checks/<int:year>/<int:month>")
def data_checks(year, month):
    last = cal_mod.monthrange(year, month)[1]
    conn = get_conn()
    results = list(_monthly_totals(conn, year, month).values())
    active_hotels = rows_to_list(conn.execute("SELECT * FROM hotels WHERE active_status=1").fetchall())
    entries = rows_to_list(conn.execute("""
        SELECT * FROM pool_stock WHERE entry_date BETWEEN ? AND ?
    """, (f"{year}-{month:02d}-01", f"{year}-{month:02d}-{last:02d}")).fetchall())
    conn.close()

    result_ids = {r["hotel_id"] for r in results}
    missing = [h["name"] for h in active_hotels if h["id"] not in result_ids]
    blank_totals = [r["hotel_name"] for r in results if r["total_income"] == 0]
    ref_map = defaultdict(list)
    for e in entries:
        if e["packing_list_ref"]: ref_map[e["packing_list_ref"]].append(e["id"])
    dup_refs = {k: v for k, v in ref_map.items() if len(v) > 1}

    issues = []
    if missing:     issues.append({"type":"missing_monthly_data","detail":f"No data for: {', '.join(missing)}"})
    if blank_totals:issues.append({"type":"blank_totals","detail":f"Zero totals: {', '.join(blank_totals)}"})
    if dup_refs:    issues.append({"type":"duplicate_refs","detail":f"{len(dup_refs)} duplicate packing list ref(s)"})

    return jsonify({
        "year": year, "month": month,
        "issue_count": len(issues), "issues": issues,
        "missing_hotels": missing,
        "blank_total_hotels": blank_totals,
        "duplicate_packing_refs": dup_refs,
    })


# ────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORTS
# ────────────────────────────────────────────────────────────────────────────

H_FILL = PatternFill("solid", fgColor="1B3A6B")
H_FONT = Font(bold=True, color="FFFFFF", size=11)
T_FONT = Font(bold=True)
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
THIN = Side(border_style="thin", color="CCCCCC")
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _aw(ws):
    for col in ws.columns:
        mx = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx+4, 42)


def _head(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = H_FILL; cell.font = H_FONT
        cell.alignment = Alignment(horizontal="center"); cell.border = BOX


@app.get("/exports/monthly-results")
def export_monthly():
    year  = request.args.get("year",  type=int)
    month = request.args.get("month", type=int)
    conn  = get_conn()
    totals = _monthly_totals(conn, year, month)
    status_rows = {r["hotel_id"]: r["status"] for r in rows_to_list(
        conn.execute("SELECT * FROM monthly_status WHERE year=? AND month=?", (year, month)).fetchall()
    )}
    conn.close()
    results = sorted(totals.values(), key=lambda r: r["hotel_name"])
    for r in results:
        r["status"] = status_rows.get(r["hotel_id"], "in_progress")

    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = f"{MONTH_NAMES[month]} {year}"
    ws.merge_cells("A1:F1")
    ws["A1"].value = f"Laundry Manager — Monthly Results: {MONTH_NAMES[month]} {year}"
    ws["A1"].font = Font(bold=True, size=14, color="1B3A6B")
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["Hotel","Guest Laundry (£)","Staff Laundry (£)","Flat Laundry (£)",
               "Total Income (£)","Status"]
    for c, h in enumerate(headers, 1): ws.cell(2, c, h)
    _head(ws, 2, len(headers))

    tg=ts=tf=0.0
    for i, r in enumerate(results, 3):
        vals = [r["hotel_name"], r["guest_laundry_income"], r["staff_laundry_income"],
                r["flat_laundry_income"], r["total_income"],
                r["status"].replace("_"," ").title()]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(i, c, v); cell.border = BOX
            if 1 < c < 6: cell.number_format = "£#,##0.00"
        tg+=r["guest_laundry_income"]; ts+=r["staff_laundry_income"]
        tf+=r["flat_laundry_income"]

    tr = len(results)+3
    for c, v in enumerate(["TOTAL", tg, ts, tf, tg+ts+tf, ""], 1):
        cell = ws.cell(tr, c, v); cell.font=T_FONT; cell.fill=SUB_FILL; cell.border=BOX
        if 1 < c < 6: cell.number_format = "£#,##0.00"

    _aw(ws)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"laundry_monthly_{year}_{month:02d}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/exports/pool-stock")
def export_pool():
    year  = request.args.get("year",  type=int)
    month = request.args.get("month", type=int)
    last  = cal_mod.monthrange(year, month)[1]
    conn  = get_conn()
    entries = rows_to_list(conn.execute("""
        SELECT ps.*, h.name as hotel_name FROM pool_stock ps
        JOIN hotels h ON h.id=ps.hotel_id
        WHERE ps.entry_date BETWEEN ? AND ? ORDER BY ps.entry_date, h.name
    """, (f"{year}-{month:02d}-01", f"{year}-{month:02d}-{last:02d}")).fetchall())
    conn.close()

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Pool Stock Detail"
    ws.merge_cells("A1:F1")
    ws["A1"].value = f"Laundry Manager — Pool Stock: {MONTH_NAMES[month]} {year}"
    ws["A1"].font = Font(bold=True, size=14, color="1B3A6B")
    ws["A1"].alignment = Alignment(horizontal="center")

    for c, h in enumerate(["Date","Hotel","Linen Item","Quantity","Packing List Ref","Notes"], 1):
        ws.cell(2, c, h)
    _head(ws, 2, 6)

    by_hotel = defaultdict(int); by_item = defaultdict(int)
    for i, e in enumerate(entries, 3):
        for c, v in enumerate([e["entry_date"],e["hotel_name"],e["linen_item"],
                                e["quantity"],e["packing_list_ref"],e["notes"]], 1):
            ws.cell(i, c, v).border = BOX
        by_hotel[e["hotel_name"]] += e["quantity"]
        by_item[e["linen_item"]]  += e["quantity"]

    ws2 = wb.create_sheet("Summary")
    ws2["A1"].value = f"Summary — {MONTH_NAMES[month]} {year}"
    ws2["A1"].font = Font(bold=True, size=13, color="1B3A6B")
    ws2.cell(3, 1, "By Hotel").font = Font(bold=True)
    for i, (n, q) in enumerate(sorted(by_hotel.items()), 4):
        ws2.cell(i, 1, n); ws2.cell(i, 2, q)
    off = len(by_hotel) + 6
    ws2.cell(off, 1, "By Linen Item").font = Font(bold=True)
    for i, (n, q) in enumerate(sorted(by_item.items()), off+1):
        ws2.cell(i, 1, n); ws2.cell(i, 2, q)
    _aw(ws); _aw(ws2)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"laundry_poolstock_{year}_{month:02d}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/exports/ranking")
def export_ranking():
    year  = request.args.get("year",  type=int)
    month = request.args.get("month", type=int)
    with app.test_request_context(f"/?year={year}&month={month}"):
        data = hotel_ranking(year, month).get_json()

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Hotel Ranking"
    ws.merge_cells("A1:H1")
    ws["A1"].value = f"Laundry Manager — Hotel Performance: {MONTH_NAMES[month]} {year}"
    ws["A1"].font = Font(bold=True, size=14, color="1B3A6B")
    ws["A1"].alignment = Alignment(horizontal="center")

    for c, h in enumerate(["Rank","Hotel","Revenue (£)","Prev Month (£)","Growth %",
                            "Pool Stock Vol.","Contribution %","Label"], 1):
        ws.cell(2, c, h)
    _head(ws, 2, 8)

    for i, r in enumerate(data["rankings"], 3):
        vals = [r["rank"],r["hotel_name"],r["total_revenue"],r["previous_revenue"],
                r["revenue_growth_pct"],r["pool_stock_volume"],r["contribution_pct"],r["label"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(i, c, v); cell.border = BOX
            if c in (3,4): cell.number_format = "£#,##0.00"
    _aw(ws)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"laundry_ranking_{year}_{month:02d}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ────────────────────────────────────────────────────────────────────────────
# BACKUP
# ────────────────────────────────────────────────────────────────────────────
#
# Backups are a manual action (a button she clicks) plus an on-screen daily
# reminder shown by the frontend in the late afternoon. There is no automatic
# timer running in the background — keeping this simple and predictable.
#
# BACKUP_DIR can be overridden via the BACKUP_DIR environment variable, e.g.
# to point at a synced cloud folder (OneDrive, Dropbox, etc.) on a real
# deployment. As shipped, it backs up into a local "backups" folder next to
# the database, which is fine for local use and for this public demo, but
# on most free hosting tiers the filesystem is not persistent — see the
# README's "Deployment" section for details.

BACKUP_DIR = os.environ.get("BACKUP_DIR", os.path.join(BASE_DIR, "backups"))
os.makedirs(BACKUP_DIR, exist_ok=True)

MAX_BACKUPS_KEPT = 60  # roughly 2-3 months of daily backups before the oldest get cleaned up


def _cleanup_old_backups():
    """Keeps the backups folder from growing forever — deletes the oldest
    files beyond MAX_BACKUPS_KEPT."""
    files = sorted(
        (f for f in os.listdir(BACKUP_DIR) if f.startswith("laundry_backup_") and f.endswith(".db")),
        reverse=True
    )
    for old_file in files[MAX_BACKUPS_KEPT:]:
        try:
            os.remove(os.path.join(BACKUP_DIR, old_file))
        except OSError:
            pass


@app.post("/backup/run")
def run_backup():
    """Copies the live database to the backup folder with a timestamped name."""
    if not os.path.exists(DB_PATH):
        return jsonify({"error": "No database file found to back up."}), 404

    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    backup_filename = f"laundry_backup_{timestamp}.db"
    backup_path = os.path.join(BACKUP_DIR, backup_filename)

    try:
        shutil.copy2(DB_PATH, backup_path)
        _cleanup_old_backups()
    except Exception as e:
        return jsonify({"error": f"Backup failed: {str(e)}"}), 500

    return jsonify({
        "ok": True,
        "filename": backup_filename,
        "backed_up_at": timestamp,
        "backup_folder": BACKUP_DIR,
    }), 201


@app.get("/backup/history")
def backup_history():
    """Lists existing backups, most recent first, so she can see when the
    last backup actually happened."""
    files = sorted(
        (f for f in os.listdir(BACKUP_DIR) if f.startswith("laundry_backup_") and f.endswith(".db")),
        reverse=True
    )
    backups = []
    for f in files[:20]:
        full_path = os.path.join(BACKUP_DIR, f)
        size_kb = round(os.path.getsize(full_path) / 1024, 1)
        backups.append({"filename": f, "size_kb": size_kb})

    last_backup_at = None
    if files:
        # Filename format: laundry_backup_YYYY-MM-DD_HHMM.db
        try:
            ts_str = files[0].replace("laundry_backup_", "").replace(".db", "")
            last_backup_at = datetime.strptime(ts_str, "%Y-%m-%d_%H%M").isoformat()
        except ValueError:
            pass

    return jsonify({
        "backups": backups,
        "total_count": len(files),
        "last_backup_at": last_backup_at,
        "backup_folder": BACKUP_DIR,
    })


# ────────────────────────────────────────────────────────────────────────────
# PETTY CASH
# ────────────────────────────────────────────────────────────────────────────

PETTY_CASH_CATEGORIES = [
    "Canteen Food", "Staff Canteen", "Staff Benefit", "Staff Uniform",
    "Postage", "Purchases", "Motor Expense", "Travel", "Misc",
    "P&S", "Cleaning", "Carpet Cleaner", "French Polisher",
    "Waste Removal", "Paint Supplies", "Advance", "Compensation",
    "Carpet Cleaners Parking/Electric", "Other",
]

CATEGORY_ACCOUNT_CODES = {
    "Postage":                          "70.99-9100",
    "Staff Benefit":                    "70.42-4000",
    "Staff Canteen":                    "70.42-4103",
    "Staff Uniform":                    "70.42-3012",
    "Purchases":                        "70.42-4100",
    "Motor Expense":                    "70.42-2900",
    "Travel":                           "70.43-2900",
    "Misc":                             "70.42-5001",
    "P&S":                              "70.42-4105",
    "Cleaning":                         "70.42-4900",
    "Carpet Cleaner":                   "70.42-4000",
    "Carpet Cleaners Parking/Electric": "70.42-4010",
    "French Polisher":                  "70.99-8993",
    "Waste Removal":                    "70.99-8993",
    "Paint Supplies":                   "70.42-7015",
    "Advance":                          "70.42-7016",
    "Compensation":                     "70.99-9210",
    "Canteen Food":                     "70.42-5100",
}


def _voucher_total(items):
    return sum(i["amount_gbp"] for i in items)


@app.get("/petty-cash/vouchers")
def list_vouchers():
    year  = request.args.get("year",  type=int)
    month = request.args.get("month", type=int)
    conn  = get_conn()
    sql   = "SELECT * FROM petty_cash_vouchers WHERE 1=1"
    params = []
    if year and month:
        last = cal_mod.monthrange(year, month)[1]
        sql += " AND voucher_date BETWEEN ? AND ?"
        params += [f"{year}-{month:02d}-01", f"{year}-{month:02d}-{last:02d}"]
    sql += " ORDER BY voucher_date DESC"
    vouchers = rows_to_list(conn.execute(sql, params).fetchall())
    for v in vouchers:
        v["items"] = rows_to_list(conn.execute(
            "SELECT * FROM petty_cash_items WHERE voucher_id=? ORDER BY id",
            (v["id"],)
        ).fetchall())
    conn.close()
    return jsonify(vouchers)


@app.post("/petty-cash/vouchers")
def create_voucher():
    d    = request.json
    conn = get_conn()
    try:
        items = d.get("items", [])
        total = sum(float(i.get("amount_gbp", 0)) for i in items)
        conn.execute("""
            INSERT INTO petty_cash_vouchers
              (voucher_date, required_for, category, passed_by, signature, notes, total_amount)
            VALUES (?,?,?,?,?,?,?)
        """, (
            d["voucher_date"], d["required_for"],
            d.get("category", "Misc"),
            d.get("passed_by", ""), d.get("signature", ""),
            d.get("notes", ""), total
        ))
        conn.commit()
        vid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        for item in items:
            conn.execute("""
                INSERT INTO petty_cash_items
                  (voucher_id, description, amount_gbp, gross, vat)
                VALUES (?,?,?,?,?)
            """, (
                vid,
                item.get("description", ""),
                float(item.get("amount_gbp", 0)),
                float(item.get("gross", 0)),
                float(item.get("vat", 0)),
            ))
        conn.commit()
        voucher = row_to_dict(conn.execute(
            "SELECT * FROM petty_cash_vouchers WHERE id=?", (vid,)).fetchone())
        voucher["items"] = rows_to_list(conn.execute(
            "SELECT * FROM petty_cash_items WHERE voucher_id=?", (vid,)).fetchall())
        conn.close()
        return jsonify(voucher), 201
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 400


@app.patch("/petty-cash/vouchers/<int:vid>")
def update_voucher(vid):
    d    = request.json
    conn = get_conn()
    try:
        items = d.get("items", [])
        total = sum(float(i.get("amount_gbp", 0)) for i in items)
        conn.execute("""
            UPDATE petty_cash_vouchers
            SET voucher_date=?, required_for=?, category=?,
                passed_by=?, signature=?, notes=?, total_amount=?
            WHERE id=?
        """, (
            d["voucher_date"], d["required_for"],
            d.get("category", "Misc"),
            d.get("passed_by", ""), d.get("signature", ""),
            d.get("notes", ""), total, vid
        ))
        conn.execute("DELETE FROM petty_cash_items WHERE voucher_id=?", (vid,))
        for item in items:
            conn.execute("""
                INSERT INTO petty_cash_items
                  (voucher_id, description, amount_gbp, gross, vat)
                VALUES (?,?,?,?,?)
            """, (
                vid,
                item.get("description", ""),
                float(item.get("amount_gbp", 0)),
                float(item.get("gross", 0)),
                float(item.get("vat", 0)),
            ))
        conn.commit()
        voucher = row_to_dict(conn.execute(
            "SELECT * FROM petty_cash_vouchers WHERE id=?", (vid,)).fetchone())
        voucher["items"] = rows_to_list(conn.execute(
            "SELECT * FROM petty_cash_items WHERE voucher_id=?", (vid,)).fetchall())
        conn.close()
        return jsonify(voucher)
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 400


@app.delete("/petty-cash/vouchers/<int:vid>")
def delete_voucher(vid):
    conn = get_conn()
    conn.execute("DELETE FROM petty_cash_items WHERE voucher_id=?", (vid,))
    conn.execute("DELETE FROM petty_cash_vouchers WHERE id=?", (vid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.get("/petty-cash/breakdown/<int:year>/<int:month>")
def petty_cash_breakdown(year, month):
    last = cal_mod.monthrange(year, month)[1]
    conn = get_conn()
    vouchers = rows_to_list(conn.execute("""
        SELECT pcv.*, GROUP_CONCAT(pci.description, ' | ') as descriptions
        FROM petty_cash_vouchers pcv
        LEFT JOIN petty_cash_items pci ON pci.voucher_id = pcv.id
        WHERE pcv.voucher_date BETWEEN ? AND ?
        GROUP BY pcv.id
        ORDER BY pcv.voucher_date
    """, (f"{year}-{month:02d}-01", f"{year}-{month:02d}-{last:02d}")).fetchall())
    conn.close()

    totals = {cat: 0.0 for cat in PETTY_CASH_CATEGORIES}
    grand  = 0.0
    for v in vouchers:
        cat = v["category"]
        if cat in totals:
            totals[cat] += v["total_amount"]
        grand += v["total_amount"]

    return jsonify({
        "year": year, "month": month,
        "grand_total": grand,
        "category_totals": totals,
        "account_codes": CATEGORY_ACCOUNT_CODES,
        "vouchers": vouchers,
    })


@app.get("/petty-cash/export/<int:year>/<int:month>")
def export_petty_cash(year, month):
    """Export petty cash breakdown as Excel matching the existing template layout."""
    with app.test_request_context(f"/?year={year}&month={month}"):
        breakdown = petty_cash_breakdown(year, month).get_json()

    conn = get_conn()
    last = cal_mod.monthrange(year, month)[1]
    vouchers = rows_to_list(conn.execute("""
        SELECT * FROM petty_cash_vouchers
        WHERE voucher_date BETWEEN ? AND ?
        ORDER BY voucher_date
    """, (f"{year}-{month:02d}-01", f"{year}-{month:02d}-{last:02d}")).fetchall())
    for v in vouchers:
        v["items"] = rows_to_list(conn.execute(
            "SELECT * FROM petty_cash_items WHERE voucher_id=? ORDER BY id", (v["id"],)).fetchall())
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Petty Cash {MONTH_NAMES[month]} {year}"

    # Title
    ws["A1"] = "LAUNDRY MANAGER"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = "PETTY CASH BREAKDOWN"
    ws["A2"].font = Font(bold=True)

    cats = PETTY_CASH_CATEGORIES
    headers = ["Description", "Gross £", "VAT £"] + cats
    header_row = 5
    for col, h in enumerate(headers, 1):
        cell = ws.cell(header_row, col, h)
        cell.fill = H_FILL; cell.font = H_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = BOX

    # Voucher rows
    row = header_row + 1
    for v in vouchers:
        for item in v["items"]:
            ws.cell(row, 1, item["description"]).border = BOX
            ws.cell(row, 2, item.get("gross") or item["amount_gbp"]).border = BOX
            ws.cell(row, 3, item.get("vat", 0)).border = BOX
            cat = v.get("category", "Misc")
            cat_col = cats.index(cat) + 4 if cat in cats else None
            if cat_col:
                ws.cell(row, cat_col, item["amount_gbp"]).border = BOX
            row += 1

    # Totals row
    row += 1
    ws.cell(row, 1, "TOTAL").font = T_FONT
    ws.cell(row, 2, breakdown["grand_total"]).font = T_FONT
    for i, cat in enumerate(cats):
        val = breakdown["category_totals"].get(cat, 0)
        cell = ws.cell(row, i + 4, val)
        cell.font = T_FONT; cell.fill = SUB_FILL
        cell.number_format = "£#,##0.00"

    # Account codes row
    row += 1
    ws.cell(row, 1, "Account Code").font = Font(italic=True, size=9, color="888888")
    for i, cat in enumerate(cats):
        code = CATEGORY_ACCOUNT_CODES.get(cat, "")
        ws.cell(row, i + 4, code).font = Font(italic=True, size=9, color="888888")

    _aw(ws)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"laundry_pettycash_{year}_{month:02d}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ────────────────────────────────────────────────────────────────────────────
FRONTEND_PATH = os.path.join(BASE_DIR, "..", "frontend", "index.html")


@app.route("/")
@app.route("/app")
def frontend():
    return send_file(FRONTEND_PATH)


@app.get("/api/demo-info")
def demo_info():
    """Lets the frontend show a 'this is a demo' banner and know whether
    OCR / sandbox-reset features are active."""
    return jsonify({
        "demo_mode": DEMO_MODE,
        "sandbox_mode": SANDBOX_MODE,
        "ocr_available": OCR_AVAILABLE,
    })


@app.post("/api/demo-reset")
def demo_reset():
    """Resets the CURRENT VISITOR's own sandbox back to the real sample
    data. This never touches the real on-disk demo data, and never affects
    any other visitor — each visitor's edits only ever exist in their own
    private in-memory copy (see db.py). Always safe to call."""
    if not SANDBOX_MODE:
        return jsonify({"error": "Sandbox mode is disabled (DEMO_MODE=false)."}), 403
    reset_sandbox(g.sandbox_id)
    return jsonify({"ok": True})


# Runs on import, so this also happens under a production WSGI server like
# gunicorn (which imports this module rather than executing __main__).
init_db()
if DEMO_MODE:
    from seed import seed_demo_data
    seed_demo_data()

if __name__ == "__main__":
    debug = os.environ.get("FLASK_DEBUG", "false").lower() in ("1", "true", "yes")
    port = int(os.environ.get("PORT", 5001))
    app.run(host="0.0.0.0", port=port, debug=debug)
